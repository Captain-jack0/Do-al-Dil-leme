import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook , load_workbook
    
urls = []
all_links = []
website_url = 'https://www.milatgazetesi.com/haber/kategori/teknoloji/?pg='
    
for i in range(0,208):
    url = website_url + str(i)
    urls.append(website_url)
 
for url in urls:
       
    # URL'den HTML içeriğini alalım
    response = requests.get(url)


    # HTTP isteği başarılıysa devam edelim
    if response.status_code == 200:
        # BeautifulSoup ile HTML içeriğini analiz edelim
        soup = BeautifulSoup(response.content, 'html.parser')

        # Tüm linkleri içeren etiketleri bulalım
        all_tags = soup.find_all('a', href=True)

        # Her bir etiket için linkleri alalım
        for tag in all_tags[24:44:1]:
            link = tag['href']
            all_links.append(link)

       



# Alınan tüm linkleri bir DataFrame'e dönüştürelim
data = {'URL': all_links}
df = pd.DataFrame(data)
df.to_excel('urllistesi.xlsx', index=False)


# Initialize a list to store URLs and paragraphs
url_list = []

n =2
num = len(all_links)
wb = load_workbook("urllistesi.xlsx")
ws = wb.active

# Take URL input from the user until the word "done" is seen
while n<num:
    url_list.append(ws.cell(n,1).value)
    n+=1
# Create a new Excel workbook and select the active sheet

# Loop through each URL in the list
for url_index, url in enumerate(url_list, start=1):
    # Fetch HTML content
    response = requests.get(url)
    html_content = response.text

    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all paragraphs
    paragraphs = soup.find_all('p')
    text = " "
    
    for paragraph in paragraphs:
        text = text + paragraph.text

    # Add "haberturk" in the first column and extracted text in the second column
    
    ws.cell(row=(url_index ), column=1, value="Milat")
    ws.cell(row=(url_index ), column=2, value=text)
    

# Save the workbook
wb.save("output.xlsx")

print("Output saved to output.xlsx")

