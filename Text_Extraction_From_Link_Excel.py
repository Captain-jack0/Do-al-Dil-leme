import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook , load_workbook

# Initialize a list to store URLs and paragraphs
url_list = []

n =1
num =int(input("Enter a number: "))
wb = load_workbook("input.xlsx")
ws = wb.active

# Take URL input from the user until the word "done" is seen
while n<=num:
    url_list.append(ws.cell(n,1).value)
    n+=1
# Create a new Excel workbook and select the active sheet

# Loop through each URL in the list
for url_index, url in enumerate(url_list, start=1):
    # Fetch HTML content
    response = requests.get(url)
    html_content = response.text

    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all paragraphs
    paragraphs = soup.find_all('p')
    text = " "
    
    for paragraph in paragraphs:
        text = text + paragraph.text

    # Add "haberturk" in the first column and extracted text in the second column
    
    ws.cell(row=(url_index ), column=1, value="Habertürk")
    ws.cell(row=(url_index ), column=2, value=text)
    

# Save the workbook
wb.save("output.xlsx")

print("Output saved to output.xlsx")

